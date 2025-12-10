"""
Dublettenkontrolle für Swisscovery-Katalog
Sucht über SRU (Search/Retrieve via URL) nach Dubletten
"""

import requests
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
import urllib.parse
import time


class DuplicateChecker:
    def __init__(self, sru_base_url):
        """
        Initialisiere den DuplicateChecker

        Args:
            sru_base_url (str): Base URL für SRU-Suche (z.B. für Swisscovery)
        """
        self.sru_base_url = sru_base_url
        self.timeout = 10
        self.delay_between_requests = 0.3  # Sekunden zwischen Anfragen

        # Namespaces für XML-Parsing
        self.namespaces = {
            'srw': 'http://www.loc.gov/zing/srw/',
            'marc': 'http://www.loc.gov/MARC21/slim'
        }

    def search_by_isbn(self, isbn):
        """
        Suche im Katalog nach ISBN

        Args:
            isbn (str): ISBN-Nummer

        Returns:
            dict: {'found': bool, 'count': int, 'records': list}
        """
        if not isbn or isbn.strip() == '':
            return {'found': False, 'count': 0, 'records': []}

        # ISBN normalisieren (Bindestriche entfernen)
        isbn_clean = isbn.replace('-', '').replace(' ', '').strip()

        # SRU-Query erstellen
        query = f'alma.isbn={isbn_clean}'

        return self._execute_sru_search(query)

    def search_by_title(self, title):
        """
        Suche im Katalog nach Titel

        Args:
            title (str): Buchtitel

        Returns:
            dict: {'found': bool, 'count': int, 'records': list}
        """
        if not title or title.strip() == '':
            return {'found': False, 'count': 0, 'records': []}

        # Titel für SRU vorbereiten (nur erste 50 Zeichen, Sonderzeichen entfernen)
        title_clean = title.strip()[:50]

        # SRU-Query erstellen
        query = f'alma.title={title_clean}'

        return self._execute_sru_search(query)

    def search_combined(self, isbn, title):
        """
        Kombinierte Suche (zuerst ISBN, dann Titel als Fallback)

        Args:
            isbn (str): ISBN-Nummer
            title (str): Buchtitel

        Returns:
            dict: {'found': bool, 'count': int, 'records': list, 'search_type': str}
        """
        # Zuerst ISBN-Suche
        if isbn and isbn.strip():
            result = self.search_by_isbn(isbn)
            if result['found']:
                result['search_type'] = 'ISBN'
                return result

        # Fallback: Titel-Suche
        if title and title.strip():
            result = self.search_by_title(title)
            result['search_type'] = 'Titel'
            return result

        return {'found': False, 'count': 0, 'records': [], 'search_type': 'Keine'}

    def _execute_sru_search(self, query):
        """
        Führe SRU-Suche durch

        Args:
            query (str): SRU-Query-String

        Returns:
            dict: Suchergebnis
        """
        try:
            # URL-Parameter
            params = {
                'version': '1.2',
                'operation': 'searchRetrieve',
                'query': query,
                'maximumRecords': '10',
                'recordSchema': 'marcxml'
            }

            # URL erstellen
            url = f"{self.sru_base_url}?{urllib.parse.urlencode(params)}"

            # HTTP-Request
            response = requests.get(url, timeout=self.timeout)
            response.raise_for_status()

            # XML parsen
            root = ET.fromstring(response.content)

            # Anzahl der Treffer
            number_of_records_elem = root.find('.//srw:numberOfRecords', self.namespaces)
            number_of_records = int(number_of_records_elem.text) if number_of_records_elem is not None else 0

            # Records extrahieren
            records = []
            record_elements = root.findall('.//srw:record', self.namespaces)

            for record_elem in record_elements:
                record_data = self._parse_marc_record(record_elem)
                if record_data:
                    records.append(record_data)

            # Delay für nächste Anfrage
            time.sleep(self.delay_between_requests)

            return {
                'found': number_of_records > 0,
                'count': number_of_records,
                'records': records
            }

        except requests.exceptions.Timeout:
            print(f"[WARNING] Timeout bei SRU-Suche: {query}")
            return {'found': False, 'count': 0, 'records': [], 'error': 'Timeout'}
        except requests.exceptions.RequestException as e:
            print(f"[WARNING] Fehler bei SRU-Suche: {e}")
            return {'found': False, 'count': 0, 'records': [], 'error': str(e)}
        except Exception as e:
            print(f"[ERROR] Unerwarteter Fehler bei SRU-Suche: {e}")
            return {'found': False, 'count': 0, 'records': [], 'error': str(e)}

    def _parse_marc_record(self, record_elem):
        """
        Parse MARC-Record aus XML

        Args:
            record_elem: XML-Element des Records

        Returns:
            dict: Extrahierte Daten
        """
        try:
            record_data = {}

            # MARC-Record finden
            marc_record = record_elem.find('.//marc:record', self.namespaces)
            if marc_record is None:
                return None

            # Titel (245$a)
            title_field = marc_record.find('.//marc:datafield[@tag="245"]/marc:subfield[@code="a"]', self.namespaces)
            record_data['title'] = title_field.text.strip() if title_field is not None else ''

            # Autor (100$a oder 700$a)
            author_field = marc_record.find('.//marc:datafield[@tag="100"]/marc:subfield[@code="a"]', self.namespaces)
            if author_field is None:
                author_field = marc_record.find('.//marc:datafield[@tag="700"]/marc:subfield[@code="a"]', self.namespaces)
            record_data['author'] = author_field.text.strip() if author_field is not None else ''

            # ISBN (020$a)
            isbn_field = marc_record.find('.//marc:datafield[@tag="020"]/marc:subfield[@code="a"]', self.namespaces)
            record_data['isbn'] = isbn_field.text.strip() if isbn_field is not None else ''

            # Verlag (264$b)
            publisher_field = marc_record.find('.//marc:datafield[@tag="264"]/marc:subfield[@code="b"]', self.namespaces)
            record_data['publisher'] = publisher_field.text.strip() if publisher_field is not None else ''

            # Jahr (264$c)
            year_field = marc_record.find('.//marc:datafield[@tag="264"]/marc:subfield[@code="c"]', self.namespaces)
            record_data['year'] = year_field.text.strip() if year_field is not None else ''

            # Trägertyp (338$a)
            carrier_field = marc_record.find('.//marc:datafield[@tag="338"]/marc:subfield[@code="a"]', self.namespaces)
            record_data['carrier'] = carrier_field.text.strip() if carrier_field is not None else ''

            return record_data

        except Exception as e:
            print(f"[WARNING] Fehler beim Parsen des MARC-Records: {e}")
            return None

    def check_excel_file_for_duplicates(self, excel_path, output_path=None):
        """
        Prüfe Excel-Datei auf Dubletten und markiere sie

        Args:
            excel_path (str): Pfad zur Excel-Datei
            output_path (str, optional): Pfad für Ausgabedatei (Standard: überschreibt Original)

        Returns:
            dict: Statistik {'total': int, 'duplicates': int, 'errors': int}
        """
        if output_path is None:
            output_path = excel_path

        # Excel-Datei laden
        workbook = load_workbook(excel_path)
        sheet = workbook.active

        # Statistik
        stats = {
            'total': 0,
            'duplicates': 0,
            'errors': 0,
            'checked': 0
        }

        # Neue Spalten hinzufügen für Dubletten-Info
        max_col = sheet.max_column
        duplicate_col = max_col + 1    # True/False
        count_col = max_col + 2        # Anzahl Treffer
        carrier_col = max_col + 3      # 338$a aus SRU
        isbn_sru_col = max_col + 4     # 020$a aus SRU

        # Header setzen
        if sheet.cell(1, duplicate_col).value is None:
            sheet.cell(1, duplicate_col, 'Dublette')
            sheet.cell(1, duplicate_col).font = Font(bold=True)

        if sheet.cell(1, count_col).value is None:
            sheet.cell(1, count_col, 'Anzahl Treffer')
            sheet.cell(1, count_col).font = Font(bold=True)

        if sheet.cell(1, carrier_col).value is None:
            sheet.cell(1, carrier_col, '338$a')
            sheet.cell(1, carrier_col).font = Font(bold=True)

        if sheet.cell(1, isbn_sru_col).value is None:
            sheet.cell(1, isbn_sru_col, '020$a')
            sheet.cell(1, isbn_sru_col).font = Font(bold=True)

        # Farben für Markierung
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        red_font = Font(color='FF0000', bold=True)

        # Durch Zeilen iterieren (ab Zeile 2, da Zeile 1 = Header)
        for row_idx in range(2, sheet.max_row + 1):
            stats['total'] += 1

            try:
                # ISBN und Titel aus der Excel-Datei holen
                # Annahme: Spalte mit ISBN ist vorhanden (muss angepasst werden)
                isbn = None
                title = None

                # Suche nach ISBN-Spalte (020$a)
                for col_idx in range(1, max_col + 1):
                    header = str(sheet.cell(1, col_idx).value).strip()
                    if '020$a' in header or 'ISBN' in header.upper():
                        isbn = sheet.cell(row_idx, col_idx).value
                        break

                # Suche nach Titel-Spalte (24510$a)
                for col_idx in range(1, max_col + 1):
                    header = str(sheet.cell(1, col_idx).value).strip()
                    if '24510$a' in header or 'TITEL' in header.upper():
                        title = sheet.cell(row_idx, col_idx).value
                        break

                # Nur suchen, wenn ISBN oder Titel vorhanden
                if isbn or title:
                    stats['checked'] += 1
                    result = self.search_combined(str(isbn) if isbn else '', str(title) if title else '')

                    if result.get('error'):
                        stats['errors'] += 1
                        sheet.cell(row_idx, duplicate_col, 'FEHLER')
                        sheet.cell(row_idx, count_col, result.get('error', ''))
                    elif result['found']:
                        stats['duplicates'] += 1
                        sheet.cell(row_idx, duplicate_col, 'JA')
                        sheet.cell(row_idx, duplicate_col).fill = yellow_fill
                        sheet.cell(row_idx, duplicate_col).font = red_font
                        sheet.cell(row_idx, count_col, result['count'])
                        sheet.cell(row_idx, count_col).fill = yellow_fill
                        # Zusatz: 338$a und 020$a aus dem ersten Treffer in eigene Spalten schreiben
                        if result['records']:
                            first_rec = result['records'][0]

                            # 338$a (carrier)
                            carrier = first_rec.get('carrier', '')
                            sheet.cell(row_idx, carrier_col, carrier)

                            # 020$a (ISBN aus SRU)
                            isbn_sru = first_rec.get('isbn', '')
                            sheet.cell(row_idx, isbn_sru_col, isbn_sru)
                    else:
                        sheet.cell(row_idx, duplicate_col, 'NEIN')
                        sheet.cell(row_idx, count_col, 0)

            except Exception as e:
                print(f"[ERROR] Fehler in Zeile {row_idx}: {e}")
                stats['errors'] += 1
                sheet.cell(row_idx, duplicate_col, 'FEHLER')

        # Datei speichern
        workbook.save(output_path)

        return stats


def check_duplicates_in_file(excel_path, sru_url, output_path=None):
    """
    Convenience-Funktion zum Prüfen von Dubletten

    Args:
        excel_path (str): Pfad zur Excel-Datei
        sru_url (str): SRU Base URL
        output_path (str, optional): Pfad für Ausgabe

    Returns:
        dict: Statistik
    """
    checker = DuplicateChecker(sru_url)
    return checker.check_excel_file_for_duplicates(excel_path, output_path)
