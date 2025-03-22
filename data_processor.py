import os
import pandas as pd
from openpyxl import Workbook
from flask import flash
from csv_loader import CSVLoader
from duplicate_checker import check_duplicates
from datetime import datetime



class DataProcessor:
    def __init__(self, paths, current_year):
        self.paths = paths
        self.current_year = current_year
        self.columns = [
            "LDR", "008", "020$a", "040$a", "040$b", "040$e", "24510$a", "24510$c",
            "264$a", "264$b", "264$c", "336$b", "336$2",
            "337$b", "337$2", "338$b", "338$2", "905$c", "905$n", "905$o", "949$v", "949$s",
            "949$x", "949$u", "949$w", "949$d", "949$z"
        ]

        # Standardwerte direkt im Skript hinterlegen
        self.default_values = {
            "LDR": "#####nam#a22004095c#4500",
            "008": "######s###########||||######|00|#||####d",
            "040$a": "CH-ZuSLS ETH",
            "040$b": "ger",
            "040$e": "rda",
            "264$a": "[s. l.]",
            "264$c": self.current_year,
            "336$b": "txt",
            "336$2": "rdacontent",
            "337$b": "n",
            "337$2": "rdamedia",
            "338$b": "nc",
            "338$2": "rdacarrier",
            "949$w": "100"
        }

        self.mapping905c = {
            "E01": "01",
            "E03": "21",
            "E05": "01",
            "E06": "01"
        }

        # Output-Ordner definieren und Dateinamen mit Zeitstempel erzeugen
        output_folder = os.path.join(os.getcwd(), "output")
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.paths["output_file"] = os.path.join(output_folder, f"Bestellliste_{timestamp}.xlsx")

        # CSVLoader initialisieren und Mappings laden
        self.csv_loader = CSVLoader(paths)
        if not self.csv_loader.load_csv_mappings():
            raise Exception("Fehler beim Laden der CSV-Mappings")

        # Mappings aus CSVLoader
        self.mapping_949v = self.csv_loader.mapping_949v
        self.mapping_949d = self.csv_loader.mapping_949d
        self.mapping_949x = self.csv_loader.mapping_949x
        self.mapping_905o = self.csv_loader.mapping_905o
        self.articles = self.csv_loader.articles
        self.sonderzeichen = self.csv_loader.sonderzeichen

    def process_files(self, saved_files):
        wb = Workbook()
        ws = wb.active
        ws.title = "Importdaten Alma"

        # Spaltenüberschriften setzen
        for col_num, column_title in enumerate(self.columns, 1):
            ws.cell(row=1, column=col_num, value=column_title)

        start_row = 2

        for file in saved_files:
            try:
                df = pd.read_excel(file).fillna('')

                for _, row in df.iterrows():
                    # Durch alle Spalten iterieren und Werte setzen
                    for column_title in self.columns:
                        if column_title in self.columns_mapping_dict():
                            # Wert aus dem Mapping holen
                            old_column = self.columns_mapping_dict()[column_title]
                            value = str(row.get(old_column, '')).strip()
                        else:
                            # Setze den Wert auf einen leeren String, wenn es kein Mapping gibt
                            value = ''

                        # Sonderzeichen ersetzen
                        for original, replacement in self.sonderzeichen.items():
                            value = value.replace(original, replacement)

                        # Bindestriche in ISBN und .0 am Ende des Wertes entfernen
                        if column_title == "020$a":
                            value = value.replace('-', '')  # Bindestriche entfernen
                            value = value.split('.')[0]      # Entfernt '.0' am Ende des Wertes

                        # Artikel werden in <<>>-Klammern gesetzt. Die Artikel befinden sich im Articles-Mapping.
                        if column_title == "24510$a":
                            title = str(value)
                            for article, formatted_article in self.articles.items():
                                if title.lower().startswith(article + ' '):
                                    title = formatted_article + ' ' + title[len(article) + 1:]
                                    break
                            value = title

                        # Wenn der Wert leer ist, den Standardwert setzen
                        value_to_set = value if value else self.default_values.get(column_title, '')

                        # Wert in Excel-Zelle schreiben
                        col_num = self.columns.index(column_title) + 1
                        ws.cell(row=start_row, column=col_num, value=value_to_set)

                    # Zusätzliche Mappings anwenden
                    self._process_905o(ws, start_row)
                    self._process_949x(ws, start_row)
                    self._process_949d(ws, start_row)
                    self._process_949v(ws, start_row)
                    self._process_905c(ws, start_row)

                    start_row += 1

            except Exception as e:
                flash(f"Fehler beim Verarbeiten der Datei {file}: {e}", 'error')
                print(f"Fehler beim Verarbeiten der Datei {file}: {e}")

        # Entferne alle Zeilen, in denen das Feld "24510$a" (Titel) leer ist.
        self._remove_empty_rows(ws)

        # 🔥 Dublettenprüfung direkt auf der geladenen Datei ausführen
        print("\n🔎 Starte Dublettenprüfung...")
        check_duplicates(wb, "Importdaten Alma")  # 👈 Direktes Weitergeben des Workbook-Objekts

        try:
            wb.save(self.paths["output_file"])
            flash("Ergebnisdatei erfolgreich gespeichert.", 'success')
            print(f"Ergebnisdatei gespeichert: {self.paths['output_file']}")
        except Exception as e:
            flash(f"Fehler beim Speichern der Ergebnisdatei: {e}", 'error')
            print(f"Fehler beim Speichern der Ergebnisdatei: {e}")

    def columns_mapping_dict(self):
        # Mapping for columns used in processing (Dictionary format for easier lookup)
        return {
            "905$n": "Bibliothek",
            "020$a": "ISBN",
            "24510$c": "Autor(en)",
            "24510$a": "Titel",
            "264$b": "Verlag",
            "949$s": "Preis Euro",
            "949$u": "Etat",
            "949$d": "Auflage/Ausgabe",
            "949$z": "Interne Bemerkung"
        }

    def _process_905o(self, ws, row):
        col_905n = self.columns.index("905$n") + 1
        col_905o = self.columns.index("905$o") + 1
        value_905n = ws.cell(row=row, column=col_905n).value
        if value_905n and value_905n in self.mapping_905o:
            ws.cell(row=row, column=col_905o, value=self.mapping_905o[value_905n])

    def _process_949x(self, ws, row):
        col_905n = self.columns.index("905$n") + 1
        col_949x = self.columns.index("949$x") + 1
        value_905n = ws.cell(row=row, column=col_905n).value
        if value_905n and value_905n in self.mapping_949x:
            ws.cell(row=row, column=col_949x, value=self.mapping_949x[value_905n])

    def _process_949d(self, ws, row):
        col_905n = self.columns.index("905$n") + 1
        col_949d = self.columns.index("949$d") + 1
        value_905n = ws.cell(row=row, column=col_905n).value
        current_value = ws.cell(row=row, column=col_949d).value or ''
        if value_905n and value_905n in self.mapping_949d:
            new_value = self.mapping_949d[value_905n]
            combined_value = f"{new_value}, {current_value}" if current_value else new_value
            ws.cell(row=row, column=col_949d, value=combined_value)

    def _process_949v(self, ws, row):
        col_264b = self.columns.index("264$b") + 1
        col_949v = self.columns.index("949$v") + 1
        value_264b = ws.cell(row=row, column=col_264b).value
        if value_264b:
            value_949v = next((v for k, v in self.mapping_949v.items() if k in str(value_264b)), '')
            ws.cell(row=row, column=col_949v, value=value_949v)

    def _process_905c(self, ws, row):
        col_905n = self.columns.index("905$n") + 1
        col_905c = self.columns.index("905$c") + 1
        value_905n = ws.cell(row=row, column=col_905n).value
        if value_905n and value_905n in self.mapping905c:
            mapped_value = self.mapping905c[value_905n]
            ws.cell(row=row, column=col_905c, value=mapped_value)
        else:
            print(f"Kein gültiges Mapping für 905$n '{value_905n}' in Zeile {row} gefunden.")

    def _remove_empty_rows(self, ws):
        """
        Entfernt alle Zeilen (außer der Kopfzeile), in denen das Feld "24510$a" (Titel) leer ist.
        """
        # Ermittle den Spaltenindex für "24510$a" (Titel)
        title_col_index = self.columns.index("24510$a") + 1  # openpyxl verwendet 1-basierte Indizes

        # Von der letzten Zeile bis zur zweiten Zeile iterieren (Kopfzeile bleibt erhalten)
        for row in range(ws.max_row, 1, -1):
            cell_value = ws.cell(row=row, column=title_col_index).value
            if cell_value is None or str(cell_value).strip() == "":
                ws.delete_rows(row)
